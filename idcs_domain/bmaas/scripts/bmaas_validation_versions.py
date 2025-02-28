import os
import re
import yaml
import json
import argparse
import subprocess
import openpyxl
from tabulate import tabulate
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

BM_VALIDATION_CONFIGMAP_NAMESPACE = "idcs-system"
KIND_DEPLOYMENT_CONTEXT="kind-idc-global"
BM_VALIDATION_CONFIGMAP_SUFFIX = "bm-validation-operator-manager-config"
EXCEL_FILENAME = "bmaas_validation_matrix.xlsx"
BASE_PATH = "./deployment/helmfile/environments"
REGIONS = ['us-staging-1', 'us-staging-3', 'us-staging-4', 'us-region-1', 'us-region-2', 'us-region-3', 'us-region-4']

def print_usage():
    """Usage instructions."""
    print("Usage Instructions:")
    print("- To check config in other ITAC environments, set the KUBECONFIG environment variable.")
    print("- Example: export KUBECONFIG=/path/to/your/kubeconfig")
    print(f"- If KUBECONFIG is not set, the script will use '{KIND_DEPLOYMENT_CONTEXT}'. If not found, it will fail.\n")


def export_to_excel(monorepo_data_table, configmap_data_table, monorepo_data_header, configmap_data_header, filename=EXCEL_FILENAME):
    """Export data to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BMaaS Validation Versions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(wrap_text=True, vertical="center")

    # Write monorepo data table headers
    for col_num, header_text in enumerate(monorepo_data_header, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Write monorepo data table
    for row_num, row_data in enumerate(monorepo_data_table, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = alignment

    last_row_monorepo_data_table = len(monorepo_data_table) + 4

    # Write configmap data table headers
    for col_num, header_text in enumerate(configmap_data_header, 1):
        cell = ws.cell(row=last_row_monorepo_data_table, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Write configmap data table
    for row_num, row_data in enumerate(configmap_data_table, last_row_monorepo_data_table + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = alignment

    # Adjust column dimensions
    for col_num in range(1, len(monorepo_data_header) + 1):
        column_letter = get_column_letter(col_num)
        column_width = max(len(str(ws[column_letter + str(row_num)].value)) for row_num in range(1, len(monorepo_data_table) + 2))
        ws.column_dimensions[column_letter].width = column_width

    for col_num in range(1, len(configmap_data_header) + 1):
        column_letter = get_column_letter(col_num)
        column_width = max(len(str(ws[column_letter + str(row_num)].value)) for row_num in range(last_row_monorepo_data_table, last_row_monorepo_data_table + len(configmap_data_table) + 1))
        ws.column_dimensions[column_letter].width = column_width

    wb.save(filename)
    full_path = os.path.abspath(filename)
    print(f"BMaaS validation package version configuration matrix has been saved to {full_path}!")

def generate_table_data(all_instance_types, data_dict, regions):
    """Generate table data from parsed information."""
    return [
        [instance_type] + [data_dict.get(instance_type, {}).get(region, "") for region in regions]
        for instance_type in sorted(all_instance_types)
    ]
    
def parse_helmfile(file_path, region, availability_zone):
    """Parse helmfile and return validation data."""
    with open(file_path, 'r') as file:
        data = yaml.safe_load(file)

    try:
        region_data = data['regions'][region]['availabilityZones'][availability_zone]['bmValidationOperator']
        enabled_instance_types = region_data['enabledInstanceTypes']
        enabled_group_instance_types = region_data['featureFlags']['enabledGroupInstanceTypes']
        validation_task_version = region_data['validationTaskVersion']
        instance_version_map = validation_task_version['instanceVersionMap']
        cluster_version_map = validation_task_version['clusterVersionMap']
    except KeyError as e:
        raise KeyError(f"Key not found in the helmfile: {e}")

    return enabled_instance_types, enabled_group_instance_types, cluster_version_map, instance_version_map

def fetch_and_process_helmfiles(helm_config_file_paths):
    """Fetch and process data from helmfiles."""
    all_instance_types = set()
    data_dict = {}
    regions = []

    for file_path in helm_config_file_paths:
        match = re.search(r'(staging|prod)-region-(us-\w+-\d+)', file_path)
        if match:
            region = match.group(2)
            availability_zone = f"{region}a"
            regions.append(region)
            enabled_instance_types, enabled_instance_group_types, cluster_version_map, instance_version_map = parse_helmfile(file_path, region, availability_zone)

            all_instance_types.update(cluster_version_map.keys(), instance_version_map.keys())
            for instance_type in cluster_version_map.keys() | instance_version_map.keys():
                data_dict.setdefault(instance_type, {})
                cluster_version = cluster_version_map.get(instance_type, "")
                instance_version = instance_version_map.get(instance_type, "")
                cluster_enabled = "(\u2714)" if instance_type in enabled_instance_group_types else ""
                instance_enabled = "(\u2714)" if instance_type in enabled_instance_types else ""
                data_dict[instance_type][region] = f"Single: {instance_version} {instance_enabled}\nCluster: {cluster_version} {cluster_enabled}"

    return regions, all_instance_types, data_dict

def monorepo_validation_versions(save=False):
    """Fetch validation versions from monorepo"""
    print("Fetching validation configuration for all regions from monorepo")
    header = ["Instance Type"]

    helm_config_file_paths = []
    for region in REGIONS:
        env = 'staging-region' if 'staging' in region else 'prod-region'
        file_path = f"{BASE_PATH}/{env}-{region}.yaml.gotmpl"
        helm_config_file_paths.append(file_path)

    regions, all_instance_types, data_dict = fetch_and_process_helmfiles(helm_config_file_paths)

    monorepo_data = generate_table_data(all_instance_types, data_dict, regions)
    headers = header + regions
    print(tabulate(monorepo_data, headers=headers, tablefmt="grid"))
    return monorepo_data, headers

def parse_configmap_output(configmap_output):
    """Parse the output of kubectl describe command."""
    patterns = {
      'instanceVersionMap': r'instanceVersionMap: (.*?)\n',
      'clusterVersionMap': r'clusterVersionMap: (.*?)\n',
      'enabledInstanceTypes': r'enabledInstanceTypes: \[(.*?)\]',
      'enabledGroupInstanceTypes': r'enabledGroupInstanceTypes: \[(.*?)\]'
    }

    parsed_data = {}
    for key, pattern in patterns.items():
      match = re.search(pattern, configmap_output, re.DOTALL)
      if match:
        if key in ['instanceVersionMap', 'clusterVersionMap']:
          parsed_data[key] = json.loads(match.group(1).replace("'", '"'))
        else:
          parsed_data[key] = match.group(1).split(", ")

    return parsed_data.get('enabledInstanceTypes', []), parsed_data.get('enabledGroupInstanceTypes', []), parsed_data.get('clusterVersionMap', {}), parsed_data.get('instanceVersionMap', {})

def configmap_validation_versions(namespace, configmapname_suffix, save=False):
    """Fetch validation versions ITAC region cluster configmap"""
    k8scfg = os.getenv('KUBECONFIG')
    if k8scfg:
        os.environ['KUBECONFIG'] = k8scfg
    else:
        try:
            subprocess.check_output(f"kubectl config use-context {KIND_DEPLOYMENT_CONTEXT}", shell=True)
        except subprocess.CalledProcessError:
            print(f"Either KUBECONFIG environment variable is not set correctly or '{KIND_DEPLOYMENT_CONTEXT}' context is not found.\nPlease set the KUBECONFIG environment variable or ensure '{KIND_DEPLOYMENT_CONTEXT}' context is available.")
            print_usage()
            exit(0)

    get_configmaps_cmd = f"kubectl get configmap -n {namespace} -o name"
    configmaps = subprocess.check_output(get_configmaps_cmd, shell=True).decode("utf-8").splitlines()
    configmap = next(cm for cm in configmaps if configmapname_suffix in cm)
    region = configmap.split("/")[1].split("-bm")[0][:-1]
    if k8scfg is None:
        k8scfg = "KUBCONFIG is not set; using local kind deployment"
    print(f"\n\n{k8scfg}")
    print(f"\nFetching validation configuration from {configmap} in region '{region}'")

    describe_cmd = f"kubectl describe {configmap} -n {namespace}"
    configmap_output = subprocess.check_output(describe_cmd, shell=True).decode("utf-8")
    enabled_instance_types, enabled_instance_group_types, cluster_version_map, instance_version_map = parse_configmap_output(configmap_output)

    enabled_instance_types = [item.strip().replace('"', '') for item in enabled_instance_types[0].split(",")]
    enabled_instance_group_types = [item.strip().replace('"', '') for item in enabled_instance_group_types[0].split(",")]
    all_instance_types = set(cluster_version_map.keys()) | set(instance_version_map.keys())
    data_dict = {}

    for instance_type in all_instance_types:
        data_dict.setdefault(instance_type, {})
        cluster_version = cluster_version_map.get(instance_type, "")
        instance_version = instance_version_map.get(instance_type, "")
        cluster_enabled = "(\u2714)" if instance_type in enabled_instance_group_types else ""
        instance_enabled = "(\u2714)" if instance_type in enabled_instance_types else ""
        data_dict[instance_type][region] = f"Single: {instance_version} {instance_enabled}\nCluster: {cluster_version} {cluster_enabled}"

    configmap_data = generate_table_data(all_instance_types, data_dict, [region])
    headers = ["Instance Type", region]
    print(tabulate(configmap_data, headers=headers, tablefmt="grid"))
    return configmap_data, headers


def main():
    parser = argparse.ArgumentParser(description="BMaaS Validation Configuration Matrix")
    parser.add_argument('--save', action='store_true', help='Save BMaaS validation version configuration matrix as Excel worksheet')
    args = parser.parse_args()
    print_usage()

    # Fetch versions from monorepo and configmap
    monorepo_data, monorepo_data_headers = monorepo_validation_versions(save=args.save)
    configmap_data, configmap_data_headers = configmap_validation_versions(BM_VALIDATION_CONFIGMAP_NAMESPACE, BM_VALIDATION_CONFIGMAP_SUFFIX, save=args.save)

    if args.save:
        # Save as .xlsx if --save flag is set
        export_to_excel(monorepo_data, configmap_data, monorepo_data_headers, configmap_data_headers)


if __name__ == "__main__":
    main()
